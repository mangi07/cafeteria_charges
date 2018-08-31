# -*- coding: utf-8 -*-
"""
Created on Thurs Aug 30, 2018

@author: Ben.Olson

TODO: Gather 
"""

import os
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill


wb = Workbook()
filepath = os.getcwd() + "/export.xlsx"
print("Exporting results to: ", filepath)

sheet = wb.active

data=[('Id','Name','Marks'),
      (1,'ABC',50),
      (2,'CDE',100)]
# append all rows
for row in data:
    sheet.append(row)


ft = Font(color = colors.WHITE)
b2 = sheet['b2']
b2.font = ft
b2.fill = PatternFill(bgColor = colors.DARKRED, fill_type = "solid")

# save file
wb.save(filepath)

