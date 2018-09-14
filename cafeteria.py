# -*- coding: utf-8 -*-
"""
Created on Tue Aug 21 08:53:06 2018

@author: Ben.Olson
"""

from openpyxl import load_workbook
import ben.parser
import ben.checker
from datetime import datetime


# Load in the workbook
wb = load_workbook('./aug18copy.xlsx')
sheet_index = 0

# Get sheet names
print("Name of sheet being analyzed: ", wb.sheetnames[sheet_index])

sheet = wb[wb.sheetnames[sheet_index]]


parser = ben.parser.Parser(datetime(2018, 8, 20), datetime(2018, 8, 31))
# cells must be traversed in this order for parser to work correctly!
for row_index in range(1, sheet.max_row+1):
    for col_index in range(1, sheet.max_column+1): 
        parser.gather_data(sheet.cell(row=row_index, column=col_index).value,
                           row_index, col_index)


accounts = parser.get_account_data()

# specify dates during which the special rules apply
min_date = datetime(2018, 8, 20)
max_date = datetime(2018, 8, 31)

checker = ben.checker.Checker(accounts, min_date, max_date)
checker.check()

parser.print_account_data()
parser.print_to_file("export.xlsx")


# TODO: coordinate with cashiers to learn menu description in their POS system